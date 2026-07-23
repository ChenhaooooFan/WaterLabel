"""
NailVesta 发货系统（两阶段工作流）
========================================================
阶段 1（生成水单）:
    Lark CSV → 选日期 → 生成水单 → 发给打 Label 的人
    分 3 类输出:
      - 深度达人水单 (NailVesta 发出, 寄给达人)
      - 客户水单     (NailVesta 发出, 寄给客户)
      - Return Label 包裹 (顾客寄回, 顾客=发件人, NailVesta=收件人)

    重要逻辑：
      如果 Return Label 包裹列打勾，这一行仍然需要生成一张「正常发货水单」；
      同时额外生成一张「Return Label 水单」。也就是同一行会产出 2 张 label 需求。

阶段 2（拣货单 & 发货单生成器，原独立程序全功能移植）:
    上传 客人水单 CSV / 深度达人单 CSV（可只传其一）
    → 拣货单 CSV（库位 × 款式 × S/M/L 汇总）
    → 发货单 PDF（按 Tracking No. 一单一页：黄底单号 / 粉底收件人 /
      地址 / 款式库位尺码数量表格 / 红字备注；同上/空白自动继承上一行）

========================================================
【2026-06 修复】Order ID 乱码问题
    根因：pd.read_csv 默认把 18 位 Order ID 推断为 float64，
          超过浮点 ~15-16 位有效精度，末几位被截断，
          导致同一订单在不同表里 Order ID 不一致。
    修复：load_lark_data 中强制 Order ID / 电话 / Tracking 列为 str，
          并统一清洗空值与 Handle 尾随空格。
"""

import io
import os
import re
from collections import OrderedDict
from datetime import datetime
from typing import Optional

import pandas as pd
import streamlit as st

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)
    from reportlab.lib.styles import ParagraphStyle
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ============================================================
# 常量
# ============================================================
SENDER_INFO = {
    "name": "NailVesta", "company": None, "phone": "5105089943",
    "country": "US", "state": "CA", "city": "Los Angeles",
    "zip": "90071",
    # 新模板把寄件人地址拆成「地址1 / 地址2」两列
    "address1": "515 S Flower St", "address2": "Floor 18 & 19, STE 1901",
}
PACKAGE_DEFAULTS = {
    "product": "USPS-1",                # B 物流产品(产品编号)，固定常量
    # Q/R/S/T + W：样例行按文本存储，保持字符串
    "length": "20", "width": "15", "height": "2",
    "weight": "0.11",                   # T 重量(kg) / W 单位净重1
    "declare_price": 5,                 # V 申报单价1
    "cn_name": "穿戴甲", "en_name": "Press-On Nails", "qty": 1,
}

# —— 水单 Excel 模板（对齐物流商最新版「客人水单」26 列 A–Z，单行表头，列名不可改）——
# 2026-07 起替换旧 36 列「[订单]-导入模板」；发件人地址合并一列，无寄件人电话/公司列
SHUIDAN_HEADERS = [
    "客户订单号", "物流产品(产品编号)",
    "发件人姓名", "发件人国家", "发件人城市", "发件人省/州", "发件人邮编", "发件人地址",
    "收件人姓名", "收件人国家", "收件人城市", "收件人省/州", "收件人邮编", "收件人电话",
    "收件人地址一", "收件人地址二",
    "长(cm)", "宽(cm)", "高(cm)", "重量(kg)",
    "配货备注1", "申报单价1", "单位净重1", "中文品名1", "英文品名1", "数量1",
]
SIZE_COL = "Size'"
RETURN_LABEL_COL = "Return Label 包裹"

# —— 深达水单 CSV 模板（对齐 USPS_Template_final，59 列，列名/顺序不可改）——
# 必填列分两类（对应模板里的喷黄/喷绿）：
#   KOL_USPS_FIXED_COLS  = 喷黄：所有行同一个值（发件人信息/包裹规格等）
#   KOL_USPS_VARIABLE_COLS = 喷绿：逐单从 Lark 深达 CSV 解析
KOL_USPS_HEADERS = [
    "Reference ID", "Reference ID 2", "Shipping Date", "Item Description",
    "Item Quantity", "Item Weight (lb)", "Item Weight (oz)", "Item Value",
    "HS Tariff #", "Country of Origin",
    "Sender First Name", "Sender Middle Initial", "Sender Last Name",
    "Sender Company/Org Name", "Sender Address Line 1", "Sender Address Line 2",
    "Sender Address Line 3", "Sender Address Town/City", "Sender State",
    "Sender Country", "Sender ZIP Code", "Sender Urbanization Code",
    "Ship From Another ZIP Code", "Sender Email", "Sender Cell Phone",
    "Recipient Country", "Recipient First Name", "Recipient Middle Initial",
    "Recipient Last Name", "Recipient Company/Org Name",
    "Recipient Address Line 1", "Recipient Address Line 2", "Recipient Address Line 3",
    "Recipient Address Town/City", "Recipient Province", "Recipient State",
    "Recipient ZIP Code", "Recipient Urbanization Code", "Recipient Phone",
    "Recipient Email", "Service Type", "Package Type",
    "Package Weight (lb)", "Package Weight (oz)", "Length", "Width", "Height",
    "Girth", "Insured Value", "Contents", "Contents Description",
    "Package Comments", "Customs Form Reference #", "License #",
    "Certificate #", "Invoice #", "Customs Form Reference # Type",
    "HAZMAT Type", "Live Animals and Perishable Goods Indicator",
]
KOL_USPS_FIXED_VALUES = {
    "Sender Company/Org Name": "NailVesta",
    "Sender Address Line 1": "515 S Flower St",
    "Sender Address Line 2": "Floor 18 & 19, STE 1901",
    "Sender Address Town/City": "Los Angeles",
    "Sender State": "CA",
    "Sender Country": "US",
    "Sender ZIP Code": "90071",
    "Sender Email": "contact@nailvesta.com",
    "Sender Cell Phone": "5105089943",
    "Recipient Country": "US",
    "Service Type": "USPS Ground Advantage",
    "Package Type": "Choose Your Own Box",
    "Package Weight (lb)": 0.25,
    "Package Weight (oz)": 4,
    "Length": 7.87,
    "Width": 5.91,
    "Height": 0.79,
}

# 需要强制按文本读取的列（防止长数字被读成 float 丢精度）
TEXT_COLUMNS = [
    "Order ID", "Phone", "手机号", "手机号 (1)",
    "查物流 Tracking No.", "打包 Tracking No.",
]

# ============================================================
# 地址解析
# ============================================================
_US_STATES_FULL = {
    "ALABAMA","ALASKA","ARIZONA","ARKANSAS","CALIFORNIA","COLORADO","CONNECTICUT",
    "DELAWARE","FLORIDA","GEORGIA","HAWAII","IDAHO","ILLINOIS","INDIANA","IOWA",
    "KANSAS","KENTUCKY","LOUISIANA","MAINE","MARYLAND","MASSACHUSETTS","MICHIGAN",
    "MINNESOTA","MISSISSIPPI","MISSOURI","MONTANA","NEBRASKA","NEVADA",
    "NEW HAMPSHIRE","NEW JERSEY","NEW MEXICO","NEW YORK","NORTH CAROLINA",
    "NORTH DAKOTA","OHIO","OKLAHOMA","OREGON","PENNSYLVANIA","RHODE ISLAND",
    "SOUTH CAROLINA","SOUTH DAKOTA","TENNESSEE","TEXAS","UTAH","VERMONT",
    "VIRGINIA","WASHINGTON","WEST VIRGINIA","WISCONSIN","WYOMING",
    "DISTRICT OF COLUMBIA","PUERTO RICO",
}
_STATE_ABBR = {"AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL",
"IN","IA","KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH",
"NJ","NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
"VA","WA","WV","WI","WY","DC","PR"}
_STATE_FULL_TO_ABBR = {
    "ALABAMA":"AL","ALASKA":"AK","ARIZONA":"AZ","ARKANSAS":"AR","CALIFORNIA":"CA",
    "COLORADO":"CO","CONNECTICUT":"CT","DELAWARE":"DE","FLORIDA":"FL","GEORGIA":"GA",
    "HAWAII":"HI","IDAHO":"ID","ILLINOIS":"IL","INDIANA":"IN","IOWA":"IA","KANSAS":"KS",
    "KENTUCKY":"KY","LOUISIANA":"LA","MAINE":"ME","MARYLAND":"MD","MASSACHUSETTS":"MA",
    "MICHIGAN":"MI","MINNESOTA":"MN","MISSISSIPPI":"MS","MISSOURI":"MO","MONTANA":"MT",
    "NEBRASKA":"NE","NEVADA":"NV","NEW HAMPSHIRE":"NH","NEW JERSEY":"NJ","NEW MEXICO":"NM",
    "NEW YORK":"NY","NORTH CAROLINA":"NC","NORTH DAKOTA":"ND","OHIO":"OH","OKLAHOMA":"OK",
    "OREGON":"OR","PENNSYLVANIA":"PA","RHODE ISLAND":"RI","SOUTH CAROLINA":"SC",
    "SOUTH DAKOTA":"SD","TENNESSEE":"TN","TEXAS":"TX","UTAH":"UT","VERMONT":"VT",
    "VIRGINIA":"VA","WASHINGTON":"WA","WEST VIRGINIA":"WV","WISCONSIN":"WI","WYOMING":"WY",
    "DISTRICT OF COLUMBIA":"DC","PUERTO RICO":"PR",
}


def state_to_abbr(state: str) -> str:
    if not state: return ""
    s = state.strip().upper()
    if s in _STATE_FULL_TO_ABBR: return _STATE_FULL_TO_ABBR[s]
    if s in _STATE_ABBR: return s
    return state if len(state) == 2 else state


def _is_zip(s): return bool(re.fullmatch(r"\d{5}(-\d{4})?", s.strip()))


def _is_phone_line(s):
    if re.match(r"^[\sa]*\(?\+?1?\)?[\s\-\.]?\(?\d{3}\)?[\s\-\.]?\d{3}[\s\-\.]?\d{4}", s.strip()):
        return True
    return bool(re.match(r"^(Tel|Phone|WhatsApp)\s*[:：]", s.strip(), re.I))


def _normalize_state(tok):
    t = tok.strip().upper()
    if t in _US_STATES_FULL: return t
    if t in _STATE_ABBR: return t
    return None


def _clean_phone(s):
    digits = re.sub(r"\D", "", s)
    if digits.startswith("1") and len(digits) == 11: digits = digits[1:]
    return digits


def _clean_street(s):
    s = re.sub(r"\(([^)]*)\)", r" \1", s)
    s = re.sub(r"#\s+", "#", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_return_label(val) -> bool:
    if val is None: return False
    if isinstance(val, bool): return val
    if isinstance(val, (int, float)):
        try:
            if pd.isna(val): return False
        except Exception:
            pass
        return float(val) >= 1.0
    if pd.isna(val): return False
    s = str(val).strip().lower()
    if not s or s == "nan" or s == "none" or s == "0" or s == "0.0":
        return False
    truthy = {"true", "1", "1.0", "yes", "y", "是", "✓", "✔", "✅",
              "勾选", "checked", "x", "t", "已选"}
    return s in truthy


def parse_free_address(text: str) -> dict:
    if not isinstance(text, str): return {}
    s = text.strip().rstrip(",").strip()
    s = s.replace("，", ",").replace("　", " ")
    s = re.sub(r"\s*\n\s*", " ", s)
    s = re.sub(r"\s+", " ", s)

    zip_matches = list(re.finditer(r"\b(\d{5})(-\d{4})?\b", s))
    zip_code = ""
    zip_match = None
    if zip_matches:
        zip_match = zip_matches[-1]
        zip_code = zip_match.group(1)

    state = ""
    state_pos = None
    if zip_match:
        zs = zip_match.start()
        ze = zip_match.end()
        ws = max(0, zs - 30)
        we = min(len(s), ze + 10)
        window = s[ws:we]
        window_upper = window.upper()
        for full, abbr in sorted(_STATE_FULL_TO_ABBR.items(), key=lambda x: -len(x[0])):
            m = re.search(r"\b" + re.escape(full) + r"\b", window_upper)
            if m:
                state = abbr
                state_pos = ws + m.start()
                break
        if not state:
            for m in re.finditer(r"\b([A-Za-z]{2})\b\.?", window):
                tok = m.group(1).upper()
                if tok in _STATE_ABBR:
                    state = tok
                    state_pos = ws + m.start()
    if not state and not zip_match:
        s_upper = s.upper()
        for full, abbr in sorted(_STATE_FULL_TO_ABBR.items(), key=lambda x: -len(x[0])):
            m = re.search(r"\b" + re.escape(full) + r"\b", s_upper)
            if m:
                state = abbr
                state_pos = m.start()
                break

    if state_pos is not None and zip_match is not None:
        cut = min(state_pos, zip_match.start())
    elif zip_match is not None:
        cut = zip_match.start()
    elif state_pos is not None:
        cut = state_pos
    else:
        cut = len(s)
    before = s[:cut].rstrip(" ,.\t\n")

    if "," in before:
        parts = [p.strip() for p in before.split(",") if p.strip()]
        if len(parts) >= 2:
            city = parts[-1]
            street = ", ".join(parts[:-1])
        else:
            street = parts[0]; city = ""
    else:
        words = before.split()
        if len(words) >= 4:
            street_kw = {"st","street","dr","drive","rd","road","ave","avenue","blvd",
                         "ln","lane","ct","court","way","cir","circle","pkwy","parkway",
                         "pl","place","hwy","highway","xing","crossing","ter","terrace",
                         "sq","square","trl","trail","plz","plaza"}
            apt_kw = {"apt","suite","ste","unit","#","no.","no"}
            last_idx = -1
            for i in range(len(words) - 1, -1, -1):
                wc = words[i].lower().rstrip(".,#").lstrip("#")
                if wc in street_kw:
                    last_idx = i; break
                if wc in apt_kw:
                    j = i + 1
                    # 单元号紧跟在 apt/unit/suite 之后，可能是数字、字母(如 "Unit C")
                    # 或两者组合(如 "12B")，不能因为它是大写字母就误判成城市名的开头
                    if j < len(words) and re.fullmatch(r"[A-Za-z0-9]{1,5}\.?#?", words[j]):
                        j += 1
                    last_idx = j - 1 if j > i + 1 else i
                    break
            # 街道关键词（Blvd/St 等）后若紧跟门牌补充（#80 / Apt 6 / Unit C / Ste 200），
            # 应并入街道，否则会被当成城市名开头（"3705 W Pico Blvd #80" 误判城市 "#80 Los Angeles"）
            while 0 <= last_idx < len(words) - 1:
                nl = words[last_idx + 1].lower().strip(".,")
                if nl.startswith("#") or re.fullmatch(r"#?\d+[a-z]?", nl):
                    last_idx += 1
                elif nl in apt_kw:
                    last_idx += 1
                    if (last_idx + 1 < len(words)
                            and re.fullmatch(r"#?[A-Za-z0-9\-]{1,6}\.?", words[last_idx + 1])):
                        last_idx += 1
                else:
                    break
            if 0 <= last_idx < len(words) - 1:
                street = " ".join(words[:last_idx + 1])
                city = " ".join(words[last_idx + 1:])
            elif len(words) >= 2:
                city = words[-1]; street = " ".join(words[:-1])
            else:
                street = before; city = ""
        else:
            street = before; city = ""

    return {
        "street": street.strip(" ,."),
        "city": city.strip(" ,.").title() if city else "",
        "state": state,
        "zip": zip_code,
    }


def _state_from_segment(seg: str):
    original = str(seg or "").strip(" ,")
    if not original:
        return None
    upper = original.upper().replace(".", "")

    if upper in _STATE_ABBR:
        return "", upper
    if upper in _STATE_FULL_TO_ABBR:
        return "", _STATE_FULL_TO_ABBR[upper]

    for full, abbr in sorted(_STATE_FULL_TO_ABBR.items(), key=lambda x: -len(x[0])):
        m = re.search(r"(?:^|\s)" + re.escape(full) + r"$", upper)
        if m:
            before = original[:m.start()].strip(" ,")
            return before, abbr

    m = re.search(r"(?:^|\s)([A-Z]{2})$", upper)
    if m and m.group(1) in _STATE_ABBR:
        before = original[:m.start()].strip(" ,")
        return before, m.group(1)

    return None


def _split_address_components(addr_lines: list, zip_code: str = "") -> dict:
    info = {
        "street": "", "street2": "", "city": "", "state": "",
        "country": "United States", "zip": zip_code or "",
    }

    cleaned = []
    for ln in addr_lines:
        if not ln:
            continue
        t = str(ln).strip()
        if re.fullmatch(r"(?i)(United States|USA|US)", t.strip(" ,")):
            continue
        if _is_zip(t):
            continue
        cleaned.append(t)

    combined = ", ".join(cleaned)
    combined = re.sub(r"\b(United States|USA|US)\b", "", combined, flags=re.I)

    zips = list(re.finditer(r"\b(\d{5})(?:-\d{4})?\b", combined))
    if zips:
        zip_match = None
        if info["zip"]:
            for z in reversed(zips):
                if z.group(1) == info["zip"]:
                    zip_match = z
                    break
        if zip_match is None and not info["zip"]:
            zip_match = zips[-1]
            info["zip"] = zip_match.group(1)
        if zip_match is not None:
            combined = (combined[:zip_match.start()] + combined[zip_match.end():]).strip(" ,")

    parts = [p.strip() for p in combined.split(",") if p.strip()]

    state_part_idx = None
    city_in_state_part = ""
    state_abbr = ""
    for idx in range(len(parts) - 1, -1, -1):
        cand = _state_from_segment(parts[idx])
        if cand:
            city_in_state_part, state_abbr = cand
            state_part_idx = idx
            break

    if state_part_idx is not None:
        info["state"] = state_abbr
        trailing_parts = parts[state_part_idx + 1:]

        if city_in_state_part:
            info["city"] = city_in_state_part
            street_parts = parts[:state_part_idx]
        else:
            if state_part_idx - 1 >= 0:
                info["city"] = parts[state_part_idx - 1]
                street_parts = parts[:state_part_idx - 1]
            else:
                street_parts = []

        if street_parts:
            info["street"] = _clean_street(", ".join(street_parts))
        if trailing_parts:
            info["street2"] = _clean_street(" ".join(trailing_parts))
        return info

    if cleaned:
        info["street"] = _clean_street(cleaned[0])
        if len(cleaned) > 1:
            info["street2"] = _clean_street(" ".join(cleaned[1:]))
    return info


def parse_shipping_info(text):
    if not isinstance(text, str):
        return {}

    text = text.replace("，", ",").replace("　", " ")

    lines = [ln.strip() for ln in text.replace("\r", "").split("\n") if ln.strip()]
    while lines and re.search(r"[一-鿿]", lines[0]) and len(lines[0]) > 15:
        lines.pop(0)
    if not lines:
        return {}

    info = {
        "name": "", "phone": "", "street": "", "street2": "",
        "city": "", "state": "", "country": "United States", "zip": "",
    }

    zip_idx = None
    for i in range(len(lines) - 1, -1, -1):
        if _is_zip(lines[i]):
            zip_idx = i
            info["zip"] = lines[i].strip()[:5]
            break
    if not info["zip"]:
        zip_matches = list(re.finditer(r"\b(\d{5})(?:-\d{4})?\b", text))
        if zip_matches:
            info["zip"] = zip_matches[-1].group(1)

    _PHONE_RE = re.compile(
        r'\(?\+?1?\)?[\s\-\.]?\(?\d{3}\)?[\s\-\.]?\d{3}[\s\-\.]?\d{4}'
    )
    phone_idx = None
    phone_suffix = ""
    for i, ln in enumerate(lines):
        if i == zip_idx:
            continue
        m = _PHONE_RE.search(ln)
        if m:
            info["phone"] = _clean_phone(m.group(0))
            phone_idx = i
            after = ln[m.end():].strip(" ,\t")
            if after:
                phone_suffix = after
            break

    if phone_idx is not None:
        m = _PHONE_RE.search(lines[phone_idx])
        before_phone = lines[phone_idx][:m.start()].strip(" ,") if m else ""
        if before_phone and not re.fullmatch(r"(?i)(United States|USA|US)", before_phone):
            info["name"] = before_phone
        else:
            for i in range(phone_idx):
                clean = re.sub(r"^Name\s*[:：]\s*", "", lines[i], flags=re.I).strip(" ,")
                if (
                    clean
                    and not _is_phone_line(clean)
                    and not _is_zip(clean)
                    and not re.fullmatch(r"(?i)(United States|USA|US)", clean)
                ):
                    info["name"] = clean
                    break

    if not info["name"]:
        for i, ln in enumerate(lines):
            clean = str(ln).strip(" ,")
            if i in ([phone_idx] if phone_idx is not None else []) or i == zip_idx:
                continue
            if (
                clean
                and not _is_phone_line(clean)
                and not _is_zip(clean)
                and not re.fullmatch(r"(?i)(United States|USA|US)", clean)
            ):
                info["name"] = clean
                break

    start = (phone_idx + 1) if phone_idx is not None else 1
    addr_lines = []
    if phone_suffix:
        addr_lines.append(phone_suffix)
    for i in range(start, len(lines)):
        if i == zip_idx:
            continue
        ln = re.sub(r"^(Address|Street|Apt|Suite)\s*[:：]\s*", "", lines[i], flags=re.I).strip()
        if ln:
            addr_lines.append(ln)

    addr = _split_address_components(addr_lines, info["zip"])
    info.update(addr)
    info["state"] = state_to_abbr(info.get("state", ""))
    info["street"] = _clean_street(info.get("street", ""))
    info["street2"] = _clean_street(info.get("street2", "")) if info.get("street2") else ""
    return info


# ============================================================
# 数据加载
# ============================================================
@st.cache_data(show_spinner=False)
def load_lark_data(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_csv(
        io.BytesIO(file_bytes),
        dtype={col: str for col in TEXT_COLUMNS},
        keep_default_na=True,
    )
    df["日期"] = df["日期"].astype(str).str.strip()

    for col in TEXT_COLUMNS + ["Handle"]:
        if col in df.columns:
            df[col] = (
                df[col].astype(str).str.strip()
                .replace({"nan": "", "None": "", "null": "", "NaN": ""})
            )
            df[col] = df[col].replace("", pd.NA)

    return df


def count_corrupted_order_ids(df: pd.DataFrame) -> int:
    """检测被 Excel 打开保存过的 CSV：Order ID 变成科学计数法(5.77E+17)，末位精度已永久丢失"""
    if "Order ID" not in df.columns:
        return 0
    s = df["Order ID"].dropna().astype(str)
    return int(s.str.contains(r"[eE]\+", regex=True).sum())


def warn_if_corrupted_order_ids(df: pd.DataFrame):
    bad = count_corrupted_order_ids(df)
    if bad > 0:
        st.error(
            f"🚨 检测到 {bad} 行 Order ID 是科学计数法（如 5.77E+17）——"
            "这份 CSV 曾被 Excel 打开并保存过，18 位订单号末尾已永久丢失，"
            "水单会打错单号、面单核对会全部对不上。"
            "请回 Lark 重新导出 CSV 直接上传，不要用 Excel 打开后另存。"
        )


def warn_if_corrupted_tracking(df: pd.DataFrame):
    """检测被 Excel 打开保存过的 CSV：打包 Tracking No. 变成科学计数法(9.4e+21)，精度已永久丢失"""
    if "打包 Tracking No." not in df.columns:
        return
    s = df["打包 Tracking No."].dropna().astype(str)
    bad = int(s.str.contains(r"[eE]\+", regex=True).sum())
    if bad > 0:
        st.error(
            f"🚨 检测到 {bad} 行「打包 Tracking No.」是科学计数法（如 9.40013E+21）——"
            "这份 CSV 曾被 Excel 打开并保存过，Tracking 号末尾已永久丢失，无法恢复，"
            "发货单 PDF 会显示这个坏值。"
            "请回 Lark 重新导出 CSV 直接上传，不要用 Excel 打开后另存。"
        )


def _clean_str(v) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (ValueError, TypeError):
        pass
    s = str(v).strip()
    return "" if s.lower() in {"nan", "none", "null"} else s


def filter_orders_for_date(df, target_date):
    is_date = df["日期"] == target_date
    has_order_id = df["Order ID"].notna() if "Order ID" in df.columns else pd.Series([False] * len(df), index=df.index)
    has_shipping = df["Shipping Info"].notna() if "Shipping Info" in df.columns else pd.Series([False] * len(df), index=df.index)
    type_a = is_date & has_order_id & has_shipping

    is_deep_kol = (df["客诉类型"] == "深度达人") if "客诉类型" in df.columns else pd.Series([False] * len(df), index=df.index)
    has_addr = df["地址"].notna() if "地址" in df.columns else pd.Series([False] * len(df), index=df.index)
    has_kol_name = (
        df["达人Name"].notna() if "达人Name" in df.columns
        else df["Handle"].notna() if "Handle" in df.columns
        else pd.Series([False] * len(df), index=df.index)
    )
    type_b = is_date & is_deep_kol & has_addr & has_kol_name

    out = df[type_a | type_b].copy()

    def make_id(r):
        oid = _clean_str(r.get("Order ID"))
        if oid:
            return oid
        handle = _clean_str(r.get("Handle"))
        name = _clean_str(r.get("达人Name"))
        if handle:
            return f"KOL-{handle}"
        if name:
            return f"KOL-{name}"
        return "KOL-未知"

    out["Order ID"] = out.apply(make_id, axis=1)
    return out


def split_orders_by_type(orders: pd.DataFrame) -> dict:
    if RETURN_LABEL_COL in orders.columns:
        is_ret = orders[RETURN_LABEL_COL].apply(is_return_label)
    else:
        is_ret = pd.Series([False] * len(orders), index=orders.index)

    is_kol = orders["客诉类型"] == "深度达人"

    return {
        "kol":               orders[ is_kol].copy(),
        "customer":          orders[~is_kol].copy(),
        "return_label_kol":  orders[ is_ret &  is_kol].copy(),
        "return_label_cust": orders[ is_ret & ~is_kol].copy(),
    }


def is_kol_order(row) -> bool:
    return _clean_str(row.get("客诉类型")) == "深度达人"


def _get_kol_name(row) -> str:
    name = _clean_str(row.get("达人Name"))
    if name:
        return name
    return _clean_str(row.get("Handle"))


def _strip_leading_name(addr: str, name: str) -> str:
    """达人地址第一行常把收件人姓名又写了一遍，解析前剥离，避免姓名混进街道。
    只做严格等值匹配（忽略大小写/空白），不会误删真实街道。"""
    if not addr or not name:
        return addr

    def norm(s):
        return re.sub(r"\s+", " ", str(s)).strip().lower()

    n = norm(name)
    if not n:
        return addr

    lines = addr.split("\n")
    while lines and norm(lines[0]) == n:          # 多行：整行等于姓名
        lines.pop(0)
    addr = "\n".join(lines).strip()

    # 单行形式："Sarah Ferguson, 179 Cape Pointe Circle, ..."
    first = addr.split("\n")[0]
    m = re.match(r"\s*(.+?)\s*,", first)
    if m and norm(m.group(1)) == n:
        rest = first[first.find(",") + 1:].strip()
        tail = addr.split("\n")[1:]
        addr = "\n".join(([rest] if rest else []) + tail).strip()
    return addr


def get_recipient_info(row) -> dict:
    if is_kol_order(row):
        addr_raw = _clean_str(row.get("地址"))
        kol_name = _get_kol_name(row)
        # 先剥离地址里重复的收件人姓名，再解析
        parsed = parse_free_address(_strip_leading_name(addr_raw, kol_name))
        # 从"手机号"列读取电话，fallback 到"手机号 (1)"列
        raw_phone = _clean_str(row.get("手机号")) or _clean_str(row.get("手机号 (1)"))
        phone = _clean_phone(raw_phone) if raw_phone else ""
        return {
            "name": kol_name, "phone": phone,
            "street": parsed.get("street", ""), "street2": "",
            "city": parsed.get("city", ""), "state": parsed.get("state", ""),
            "country": "United States", "zip": parsed.get("zip", ""),
        }
    else:
        ship = parse_shipping_info(row.get("Shipping Info", ""))
        return {
            "name": ship.get("name", ""), "phone": ship.get("phone", ""),
            "street": ship.get("street", ""), "street2": ship.get("street2", ""),
            "city": ship.get("city", ""), "state": ship.get("state", ""),
            "country": ship.get("country", "United States"), "zip": ship.get("zip", ""),
        }


# ============================================================
# 地址完整性校验
# ============================================================
def _is_blank_value(v) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (ValueError, TypeError):
        pass
    s = str(v).strip()
    return s == "" or s.lower() in {"nan", "none", "null"}


def _valid_us_zip(v) -> bool:
    if _is_blank_value(v):
        return False
    return bool(re.fullmatch(r"\d{5}(-\d{4})?", str(v).strip()))


def _valid_us_state(v) -> bool:
    if _is_blank_value(v):
        return False
    return state_to_abbr(str(v).strip()).upper() in _STATE_ABBR


def _format_original_address(row) -> str:
    if is_kol_order(row):
        return _clean_str(row.get("地址"))
    return _clean_str(row.get("Shipping Info"))


def validate_address_info(row, label_kind: str = "正常发货") -> list:
    info = get_recipient_info(row)
    # Return Label 里 get_recipient_info 拿到的是"寄件人"（顾客/达人寄回给我们）；
    # 正常发货 / 深达 USPS 里拿到的都是"收件人"
    role = "发件人" if label_kind == "Return Label" else "收件人"
    issues = []

    def add_issue(field, message, level="严重"):
        issues.append({
            "影响Label": label_kind,
            "对象": role,
            "订单号": _clean_str(row.get("Order ID")),
            "客诉类型": _clean_str(row.get("客诉类型")),
            "问题等级": level,
            "问题字段": field,
            "问题说明": message,
            "解析姓名": str(info.get("name", "") or "").strip(),
            "解析电话": str(info.get("phone", "") or "").strip(),
            "解析地址一": str(info.get("street", "") or "").strip(),
            "解析地址二": str(info.get("street2", "") or "").strip(),
            "解析城市": str(info.get("city", "") or "").strip(),
            "解析州": state_to_abbr(str(info.get("state", "") or "").strip()),
            "解析邮编": str(info.get("zip", "") or "").strip(),
            "原始地址信息": _format_original_address(row),
        })

    raw_addr = _format_original_address(row)
    if _is_blank_value(raw_addr):
        add_issue("原始地址", "原始地址信息为空，无法生成可用 Label")
        return issues

    if _is_blank_value(info.get("name")):
        add_issue(f"{role}姓名", f"{role}姓名缺失或没有被程序识别出来")

    phone = str(info.get("phone", "") or "").strip()
    if _is_blank_value(phone):
        if is_kol_order(row):
            add_issue(f"{role}电话", f"{role}电话缺失；深度达人单可人工确认是否允许为空", level="提醒")
        else:
            add_issue(f"{role}电话", f"{role}电话缺失")
    elif not re.fullmatch(r"\d{10}", phone):
        add_issue(f"{role}电话", f"{role}电话格式异常：{phone}，建议为 10 位美国电话", level="提醒")

    if _is_blank_value(info.get("street")):
        add_issue(f"{role}地址一", f"{role}街道地址 Address 1 缺失")
    if _is_blank_value(info.get("city")):
        add_issue(f"{role}城市", f"{role}城市 City 缺失或没有被程序识别出来")

    state = str(info.get("state", "") or "").strip()
    if _is_blank_value(state):
        add_issue(f"{role}州", f"{role}州 State 缺失或没有被程序识别出来")
    elif not _valid_us_state(state):
        add_issue(f"{role}州", f"{role}州 State 格式异常：{state}，应为美国州缩写或州全称")

    zip_code = str(info.get("zip", "") or "").strip()
    if _is_blank_value(zip_code):
        add_issue(f"{role}邮编", f"{role}邮编 Zip Code 缺失或没有被程序识别出来")
    elif not _valid_us_zip(zip_code):
        add_issue(f"{role}邮编", f"{role}邮编 Zip Code 格式异常：{zip_code}，应为 5 位或 ZIP+4")

    if str(info.get("name", "") or "").strip().lower() in {"united states", "usa", "us"}:
        add_issue(f"{role}姓名", f"{role}姓名被误识别为国家，请检查原始 Shipping Info")

    return issues


def build_address_issue_report(orders: pd.DataFrame, label_kind: str = "正常发货") -> pd.DataFrame:
    all_issues = []
    if orders is None or len(orders) == 0:
        return pd.DataFrame(columns=[
            "影响Label", "对象", "订单号", "客诉类型", "问题等级", "问题字段",
            "问题说明", "解析姓名", "解析电话", "解析地址一", "解析地址二",
            "解析城市", "解析州", "解析邮编", "原始地址信息",
        ])
    for _, row in orders.iterrows():
        all_issues.extend(validate_address_info(row, label_kind=label_kind))
    return pd.DataFrame(all_issues)


def address_report_csv(report: pd.DataFrame) -> bytes:
    return report.to_csv(index=False).encode("utf-8-sig")

# ============================================================
# 文件 3：水单 Excel（正常发货：NailVesta → 客户）
# ============================================================
def _combine_address(street: str, street2: str) -> str:
    street = (street or "").strip()
    street2 = (street2 or "").strip()
    if street and street2:
        return f"{street}, {street2}"
    return street or street2


# 黄色高亮列（1-indexed，对齐最新模板样例中填黄底 FFFF00 的列）
# = 每行固定值块：B 物流产品 + C–H（发件人整段）+ J 收件人国家
#   + Q–T（长宽高重）+ V–Z（申报单价/单位净重/品名/数量）
# 高亮按「列位置」而非内容，正常单与退货单收/发件人互换也不变 → 两者共用同一套
SHUIDAN_HL_COLS = (
    set(range(2, 9))        # B–H
    | {10}                  # J
    | set(range(17, 21))    # Q–T
    | set(range(22, 27))    # V–Z
)
HL_FILL_RGB = "FFFF00"


def _write_shuidan_workbook(rows_data: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    hl_fill = PatternFill("solid", fgColor=HL_FILL_RGB)
    header_font = Font(name="宋体", size=11, bold=True)
    hl_font = Font(name="Calibri", size=11)
    plain_font = Font(name="宋体", size=10)
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"

    # 第 1 行：列头（单行表头，列名不可改；宋体11 粗体居中，无底色）
    for c, h in enumerate(SHUIDAN_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.alignment = center

    # 第 2 行起：数据。固定值列黄底 Calibri11，逐单列 宋体10
    for r_idx, row_vals in enumerate(rows_data, 2):
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=val)
            if c in SHUIDAN_HL_COLS:
                cell.font = hl_font
                cell.fill = hl_fill
            else:
                cell.font = plain_font

    # 防止长订单号/电话/邮编被 Excel 转成科学计数法或丢掉前导零
    # A 客户订单号 / G 发件人邮编 / M 收件人邮编 / N 收件人电话
    for col_idx in (1, 7, 13, 14):
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = "@"

    # 列宽对齐样例：A–Z 全部 16
    for c in range(1, len(SHUIDAN_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _order_no(r) -> str:
    """A 客户单号：真实 Order ID 才输出；深达单占位号 KOL-xxx 按用户要求留空"""
    oid = _clean_str(r.get("Order ID"))
    return "" if oid.startswith("KOL-") else oid


def _package_tail() -> list:
    """Q–Z：长宽高重 + 配货备注1(空) + 申报单价/单位净重/品名/数量，全行同值"""
    return [
        PACKAGE_DEFAULTS["length"], PACKAGE_DEFAULTS["width"],
        PACKAGE_DEFAULTS["height"], PACKAGE_DEFAULTS["weight"],
        None, PACKAGE_DEFAULTS["declare_price"], PACKAGE_DEFAULTS["weight"],
        PACKAGE_DEFAULTS["cn_name"], PACKAGE_DEFAULTS["en_name"], PACKAGE_DEFAULTS["qty"],
    ]


def build_shuidan_xlsx(orders: pd.DataFrame) -> bytes:
    """正常发货：发件人 = NailVesta，收件人 = 客户/达人"""
    sender_addr = f'{SENDER_INFO["address1"]}, {SENDER_INFO["address2"]}'
    rows_data = []
    for _, r in orders.iterrows():
        recip = get_recipient_info(r)
        rows_data.append([
            _order_no(r), PACKAGE_DEFAULTS["product"],
            SENDER_INFO["name"], SENDER_INFO["country"], SENDER_INFO["city"],
            SENDER_INFO["state"], SENDER_INFO["zip"], sender_addr,
            recip["name"], "US", recip["city"], state_to_abbr(recip["state"]),
            recip["zip"], recip["phone"],
            recip["street"], recip.get("street2") or None,
        ] + _package_tail())
    return _write_shuidan_workbook(rows_data)


def build_return_label_xlsx(orders: pd.DataFrame) -> bytes:
    """Return Label 寄回：发件人 = 客户/达人，收件人 = NailVesta
    发件人地址只有一列 → 达人 street/street2 合并；收件人地址一/二 = NailVesta 地址1/2"""
    rows_data = []
    for _, r in orders.iterrows():
        sender = get_recipient_info(r)
        addr = sender["street"]
        if sender.get("street2"):
            addr = f'{addr}, {sender["street2"]}'
        rows_data.append([
            _order_no(r), PACKAGE_DEFAULTS["product"],
            sender["name"], "US", sender["city"], state_to_abbr(sender["state"]),
            sender["zip"], addr,
            SENDER_INFO["name"], SENDER_INFO["country"], SENDER_INFO["city"],
            SENDER_INFO["state"], SENDER_INFO["zip"], SENDER_INFO["phone"],
            SENDER_INFO["address1"], SENDER_INFO["address2"],
        ] + _package_tail())
    return _write_shuidan_workbook(rows_data)


def _split_kol_name(full_name: str) -> tuple:
    """USPS 模板要拆 First/Last Name；取第一个词做 First，剩余做 Last。"""
    parts = str(full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def build_kol_usps_csv(orders: pd.DataFrame, ship_date_str: str) -> bytes:
    """深达水单：对齐 USPS_Template_final 59 列，喷黄=固定值，喷绿=逐单解析"""
    try:
        ship_date = datetime.strptime(ship_date_str, "%Y/%m/%d").strftime("%m/%d/%Y")
    except ValueError:
        ship_date = ship_date_str

    rows = []
    for _, r in orders.iterrows():
        recip = get_recipient_info(r)
        first, last = _split_kol_name(recip["name"])
        row = {h: None for h in KOL_USPS_HEADERS}
        row.update(KOL_USPS_FIXED_VALUES)
        row["Shipping Date"] = ship_date
        row["Recipient First Name"] = first
        row["Recipient Last Name"] = last
        row["Recipient Address Line 1"] = recip["street"]
        row["Recipient Address Town/City"] = recip["city"]
        row["Recipient State"] = state_to_abbr(recip["state"])
        row["Recipient ZIP Code"] = recip["zip"]
        row["Recipient Phone"] = recip["phone"]
        rows.append(row)

    df = pd.DataFrame(rows, columns=KOL_USPS_HEADERS)
    return df.to_csv(index=False).encode("utf-8-sig")


# ============================================================================
# 拣货单 & 发货单生成器（原程序全功能移植，逻辑一字未改）
# ============================================================================
def _register_cjk_font() -> str:
    """Try to register a TTF font that covers all CJK (incl. traditional).
    Falls back to the built-in CID font if none found."""
    candidates = [
        # Linux / Streamlit Cloud (install via packages.txt: fonts-noto-cjk)
        "/usr/share/fonts/truetype/noto/NotoSansCJKsc-Regular.otf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        # macOS
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/Library/Fonts/Songti.ttc",
        "/System/Library/Fonts/Supplemental/Songti.ttc",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("CJKFont", path))
                return "CJKFont"
            except Exception:
                continue
    # Built-in CID fallback (covers GB2312 simplified CJK)
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    return "STSong-Light"

CJK = _register_cjk_font() if REPORTLAB_OK else None


# ── Helpers ──────────────────────────────────────────────────────────────────

def extract_location(loc_field: str) -> str:
    m = re.search(r"库位：(\S+)", str(loc_field))
    return m.group(1).strip() if m else str(loc_field).strip()


# 没有固定库位的杂项（Organizer Binder / Toolkits / Storage Box 等）在主表里
# 库位字段本来就是空的，原正则要求 \S+（至少一个非空白字符）导致整条商品被 continue 跳过、
# 完全不出现在拣货单/发货单里。用这个占位符替代"直接丢弃"。
NO_LOCATION_LABEL = "无固定库位·人工核对"


def split_shipping_info(raw: str) -> tuple:
    """Return (name, address) from multi-line Shipping Info.
    Format: Name / (+1)Phone / Street / City,State,Country / Zip
    """
    lines = [l.strip() for l in str(raw).split("\n") if l.strip()]
    if not lines:
        return "", ""
    name = lines[0]
    # Drop the phone line (starts with parenthesis)
    addr_lines = [l for l in lines[1:] if not re.match(r"^\(", l)]
    return name, "\n".join(addr_lines)


def is_blank(val: str) -> bool:
    return val.strip().lower() in ("", "nan")


# ── Parsers ──────────────────────────────────────────────────────────────────

def parse_customer_rows(df: pd.DataFrame, include_no_location: bool = True) -> tuple:
    rows = []
    skipped = []  # [(收件人, 商品名)] 无固定库位且用户选择"不添加"时，记录下来提醒人工核对
    prev_tracking = prev_name = prev_address = prev_note = ""
    order_seq = 0  # 每遇到一个新地址块（非"同上"/非空）就是新订单，避免多单因 Tracking 同为空而被误合并成一页

    for _, r in df.iterrows():
        size         = str(r.get("Size'", "")).strip()
        loc_raw      = str(r.get("库位", "")).strip()
        tracking_raw = str(r.get("打包 Tracking No.", "")).strip()
        note         = str(r.get("備註", "")).strip()
        shipping_raw = str(r.get("Shipping Info", "")).strip()

        # 同上 or blank shipping info → inherit name/address（先算这个，用来判断是否新订单）
        is_new_order = not (is_blank(shipping_raw) or shipping_raw == "同上")
        if is_new_order:
            name, address = split_shipping_info(shipping_raw)
            prev_name, prev_address = name, address
            order_seq += 1
        else:
            name, address = prev_name, prev_address

        # 同上 or blank tracking → inherit from previous row；仍为空则保留空，不臆造
        if is_blank(tracking_raw) or tracking_raw == "同上":
            tracking = prev_tracking
        else:
            tracking = tracking_raw
            prev_tracking = tracking

        if is_blank(note):
            note = ""

        # Parse product name + location directly from 库位 column
        # Each entry format: "Product Name ｜ 库位：A-01-01"
        for part in loc_raw.split(","):
            part = part.strip()
            if not part:
                continue
            m = re.match(r"(.+?)\s*｜\s*库位：(\S+)", part)
            if m:
                product = m.group(1).strip()
                loc     = m.group(2).strip()
            else:
                # 商品名能解析出来，但"库位："后面是空的（无固定库位的杂项）
                m2 = re.match(r"(.+?)\s*｜\s*库位：\s*$", part)
                if not m2:
                    continue  # 真正的格式错误，跳过
                product = m2.group(1).strip()
                if not include_no_location:
                    skipped.append((name, product))
                    continue
                loc = NO_LOCATION_LABEL
            rows.append(dict(tracking=tracking, order_key=order_seq, name=name, address=address,
                             product=product, location=loc, size=size, note=note))
    return rows, skipped


def parse_influencer_rows(df: pd.DataFrame, include_no_location: bool = True) -> tuple:
    rows = []
    skipped = []  # [(达人Name, 商品名)]
    for idx, (_, r) in enumerate(df.iterrows()):
        size         = str(r.get("Size", "")).strip()
        loc_raw      = str(r.get("款式 + 库位", "")).strip()
        tracking     = str(r.get("打包 Tracking No.", "")).strip()
        note         = str(r.get("備註", "")).strip()
        name         = str(r.get("达人Name", "")).strip()
        address      = str(r.get("地址", "")).strip()

        if is_blank(tracking):
            tracking = ""
        if is_blank(note):
            note = ""

        entries = [e.strip() for e in loc_raw.split(",") if e.strip()]
        for entry in entries:
            m = re.match(r"(.+?)\s*｜\s*库位：(\S+)", entry)
            if m:
                product, loc = m.group(1).strip(), m.group(2).strip()
            else:
                m2 = re.match(r"(.+?)\s*｜\s*库位：\s*$", entry)
                product = m2.group(1).strip() if m2 else entry
                if not include_no_location:
                    skipped.append((name, product))
                    continue
                loc = NO_LOCATION_LABEL
            # 每行 iterrows 本身就是一个独立订单，用行号做 order_key，Tracking 缺失也不会互相合并
            rows.append(dict(tracking=tracking, order_key=idx, name=name, address=address,
                             product=product, location=loc, size=size, note=note))
    return rows, skipped


# ── Pick list (CSV) ──────────────────────────────────────────────────────────

def make_pick_list(rows: list) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["库位", "款式", "S", "M", "L", "合计"])
    df = pd.DataFrame(rows)
    df["qty"] = 1
    pivot = (df.groupby(["location", "product", "size"])["qty"]
               .sum().unstack(fill_value=0).reset_index())
    pivot.columns.name = None
    for col in ["S", "M", "L"]:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot["合计"] = pivot[["S", "M", "L"]].sum(axis=1)
    pivot = pivot[["location", "product", "S", "M", "L", "合计"]]
    pivot.columns = ["库位", "款式", "S", "M", "L", "合计"]
    pivot["_no_loc"] = pivot["库位"].eq(NO_LOCATION_LABEL)
    return (pivot.sort_values(["_no_loc", "库位", "款式"])
                 .drop(columns="_no_loc").reset_index(drop=True))


# ── Shipping list (PDF) ──────────────────────────────────────────────────────

def make_shipping_pdf(rows: list) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=0.6*inch, rightMargin=0.6*inch,
                            topMargin=0.6*inch, bottomMargin=0.6*inch)

    YELLOW     = colors.HexColor("#FFE600")
    PINK       = colors.HexColor("#FFD6E0")
    HEADER_BG  = colors.HexColor("#4A90D9")

    tracking_style = ParagraphStyle(
        "tracking", fontName=CJK, fontSize=22, leading=28,
        spaceAfter=10, backColor=YELLOW,
        leftIndent=6, rightIndent=6,
    )
    name_style = ParagraphStyle(
        "name", fontName=CJK, fontSize=16, leading=22,
        spaceAfter=4, backColor=PINK,
        leftIndent=6, rightIndent=6,
    )
    addr_style = ParagraphStyle(
        "addr", fontName=CJK, fontSize=14, leading=20,
        spaceAfter=3, leftIndent=6,
    )
    note_style = ParagraphStyle(
        "note", fontName=CJK, fontSize=14, leading=20,
        spaceAfter=3, textColor=colors.red, leftIndent=6,
    )

    # Group rows preserving insertion order。按 order_key（订单）分组，
    # 而不是按 tracking 值分组——避免多个"暂无 Tracking"的不同订单被误合并到同一页
    groups = OrderedDict()
    for row in rows:
        groups.setdefault(row["order_key"], []).append(row)

    story = []
    first_page = True

    for order_key, items in groups.items():
        if not first_page:
            story.append(PageBreak())
        first_page = False

        tracking = items[0]["tracking"]
        # 注意：CJK 字体不一定覆盖 emoji glyph，用纯文字提示，避免 PDF 里出现乱码/空字符
        tracking_display = tracking if tracking else "暂无 Tracking No.（请先完成打包）"

        # ── Tracking number (yellow highlight) ──
        story.append(Paragraph(f"Tracking No.: {tracking_display}", tracking_style))
        story.append(Spacer(1, 0.12*inch))

        # ── Recipient name (pink highlight) + address ──
        name    = items[0]["name"]
        address = items[0]["address"]

        story.append(Paragraph(name, name_style))
        for line in address.split("\n"):
            if line.strip():
                story.append(Paragraph(line.strip(), addr_style))
        story.append(Spacer(1, 0.2*inch))

        # ── Product table (aggregate qty) ──
        merged = {}
        for item in items:
            key = (item["product"], item["location"], item["size"])
            merged[key] = merged.get(key, 0) + 1

        table_data = [["款式", "库位", "尺码", "数量"]]
        for (product, loc, sz), qty in merged.items():
            table_data.append([product, loc, sz, str(qty)])

        usable = 7.3 * inch  # letter width minus margins
        col_widths = [usable * 0.50, usable * 0.25, usable * 0.13, usable * 0.12]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME",       (0, 0), (-1, -1), CJK),
            ("FONTSIZE",       (0, 0), (-1,  0), 16),   # header row
            ("FONTSIZE",       (0, 1), (-1, -1), 15),   # data rows
            ("BACKGROUND",     (0, 0), (-1,  0), HEADER_BG),
            ("TEXTCOLOR",      (0, 0), (-1,  0), colors.white),
            ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",          (0, 1), (0,  -1), "LEFT"),
            ("GRID",           (0, 0), (-1, -1), 0.8, colors.HexColor("#AAAAAA")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F0F4F8")]),
            ("TOPPADDING",     (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 10),
            ("LEFTPADDING",    (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
        ]))
        story.append(tbl)

        # ── Notes ──
        all_notes = list(dict.fromkeys(
            item["note"] for item in items if item["note"]
        ))
        if all_notes:
            story.append(Spacer(1, 0.15*inch))
            for n in all_notes:
                story.append(Paragraph(f"备注：{n}", note_style))

    doc.build(story)
    return buf.getvalue()


# ── Utilities ────────────────────────────────────────────────────────────────

def to_csv_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue()


def extract_date_prefix(df: pd.DataFrame) -> str:
    """Return MMDD string from the first non-empty 日期 value in the dataframe."""
    try:
        raw = df["日期"].dropna().iloc[0]
        dt = pd.to_datetime(raw)
        return dt.strftime("%m%d")
    except Exception:
        return ""


def show_section(label: str, rows: list, key_prefix: str, date_prefix: str):
    pick = make_pick_list(rows)
    pick_name = f"{date_prefix}{label}水单拣货单.csv"
    ship_name = f"{date_prefix}{label}水单发货单.pdf"

    st.subheader(f"{label} — 拣货单")
    st.dataframe(pick, use_container_width=True, hide_index=True)
    st.download_button(
        f"⬇ 下载{label}拣货单 CSV",
        to_csv_bytes(pick),
        pick_name,
        "text/csv",
        key=f"{key_prefix}_pick",
    )

    st.divider()
    st.subheader(f"{label} — 发货单")
    pdf_bytes = make_shipping_pdf(rows)
    st.download_button(
        f"⬇ 下载{label}发货单 PDF",
        pdf_bytes,
        ship_name,
        "application/pdf",
        key=f"{key_prefix}_ship",
    )


# 长数字列必须按文本读取，否则 pandas 会推断成 float64，
# 18-22 位的 Tracking No. / Order ID 超出浮点精度会变成科学计数法（如 9.4e+21）
_ID_LIKE_COLS = ["打包 Tracking No.", "查物流 Tracking No.", "Order ID",
                 "Phone", "手机号", "手机号 (1)"]


def read_file(f) -> pd.DataFrame:
    name = f.name.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(f, dtype={c: str for c in _ID_LIKE_COLS})
    else:
        df = pd.read_csv(f, dtype={c: str for c in _ID_LIKE_COLS})
    for c in _ID_LIKE_COLS:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip().str.replace(
                r"\.0$", "", regex=True
            ).replace({"nan": "", "None": ""})
    return df

# ============================================================
# Streamlit UI
# ============================================================
st.set_page_config(page_title="NailVesta 发货系统", page_icon="💅", layout="wide")
st.title("💅 NailVesta 发货系统")

tab1, tab2 = st.tabs([
    "1️⃣ 第一步：生成水单（发给打 Label 的人）",
    "2️⃣ 第二步：生成拣货单 CSV + 发货单 PDF",
])

# ============================================================
# Tab 1: 生成水单
# ============================================================
with tab1:
    st.subheader("第 1 步 · 生成水单 Excel")
    st.caption(
        "上传 Lark CSV → 选日期 → 下载水单（正常发货 + 额外 Return Label）→ 发给打 Label 的同事"
    )

    lark_file_1 = st.file_uploader("Lark 水单 CSV", type=["csv"], key="lark1")

    if lark_file_1 is None:
        st.info("👆 请上传 Lark CSV")
    else:
        df1 = load_lark_data(lark_file_1.read())
        warn_if_corrupted_order_ids(df1)
        all_dates_1 = sorted(
            [d for d in df1["日期"].dropna().unique()
             if re.match(r"^\d{4}/\d{1,2}/\d{1,2}$", str(d))],
            key=lambda x: datetime.strptime(x, "%Y/%m/%d"), reverse=True,
        )
        if not all_dates_1:
            st.error("CSV 中找不到合法日期")
        else:
            sel_date_1 = st.selectbox(
                "发货日期（洛杉矶时间，默认最晚）",
                options=all_dates_1, index=0,
                format_func=lambda x: f"{x} {'⬅️ 最新' if x == all_dates_1[0] else ''}",
                key="d1",
            )
            orders_1 = filter_orders_for_date(df1, sel_date_1)
            split_1 = split_orders_by_type(orders_1)
            kol_orders = split_1["kol"]
            customer_orders = split_1["customer"]
            return_kol_orders = split_1["return_label_kol"]
            return_cust_orders = split_1["return_label_cust"]
            return_total = len(return_kol_orders) + len(return_cust_orders)

            mc1, mc2, mc3, mc4, mc5 = st.columns(5)
            mc1.metric("当日总订单", len(orders_1))
            mc2.metric("深度达人单", len(kol_orders))
            mc3.metric("客户单", len(customer_orders))
            mc4.metric("↩️ 退货达人", len(return_kol_orders))
            mc5.metric("↩️ 退货客户", len(return_cust_orders))

            if len(orders_1) == 0:
                st.warning(f"⚠️ {sel_date_1} 没有可发货订单")
            else:
                with st.expander("📊 客诉类型分布", expanded=False):
                    type_counts = orders_1["客诉类型"].fillna("(空)").value_counts()
                    st.dataframe(
                        type_counts.rename_axis("客诉类型").reset_index(name="单数"),
                        use_container_width=True, hide_index=True,
                    )

                with st.expander("👀 订单预览", expanded=False):
                    preview_cols = ["Order ID", "客诉类型", RETURN_LABEL_COL,
                                    "Product Name", "款式",
                                    "Full SKU", SIZE_COL, "Size",
                                    "达人Name", "Handle", "地址", "Shipping Info"]
                    preview = orders_1[[c for c in preview_cols if c in orders_1.columns]]
                    st.dataframe(preview, use_container_width=True, height=300)

                date_compact = sel_date_1.replace("/", "")

                # —— 地址完整性校验 ——
                normal_orders_for_check = pd.concat([kol_orders, customer_orders], ignore_index=False)
                normal_address_issues = build_address_issue_report(normal_orders_for_check, label_kind="正常发货")
                return_address_issues = pd.concat([
                    build_address_issue_report(return_kol_orders, label_kind="Return Label"),
                    build_address_issue_report(return_cust_orders, label_kind="Return Label"),
                ], ignore_index=True)
                address_issues_all = pd.concat(
                    [normal_address_issues, return_address_issues],
                    ignore_index=True,
                )

                if len(address_issues_all) > 0:
                    serious_cnt = (address_issues_all["问题等级"] == "严重").sum()
                    remind_cnt = (address_issues_all["问题等级"] == "提醒").sum()
                    st.warning(
                        f"⚠️ 地址信息检查发现 {len(address_issues_all)} 条问题："
                        f"{serious_cnt} 条严重 / {remind_cnt} 条提醒。"
                        "建议先修正 Lark 水单表后再下载或上传水单。"
                    )
                    with st.expander("🚨 查看地址问题明细", expanded=True):
                        show_cols = [
                            "影响Label", "订单号", "客诉类型", "问题等级", "问题字段", "问题说明",
                            "解析姓名", "解析电话", "解析地址一", "解析地址二",
                            "解析城市", "解析州", "解析邮编", "原始地址信息",
                        ]
                        st.dataframe(
                            address_issues_all[[c for c in show_cols if c in address_issues_all.columns]],
                            use_container_width=True, height=320, hide_index=True,
                        )
                        st.download_button(
                            f"📥 下载地址问题报告（{len(address_issues_all)} 条）",
                            data=address_report_csv(address_issues_all),
                            file_name=f"{date_compact}地址问题报告.csv",
                            mime="text/csv",
                            use_container_width=True,
                            key="dl_address_issues_tab1",
                        )
                else:
                    st.success("✅ 地址信息检查通过：姓名 / 电话 / 地址 / 城市 / 州 / 邮编未发现明显缺失。")

                st.divider()
                st.markdown("### 📥 下载水单文件（正常发货）")

                col_kol, col_other = st.columns(2)

                with col_kol:
                    st.markdown("#### 🌟 深度达人水单")
                    if len(kol_orders) == 0:
                        st.info("今日无深度达人单")
                    else:
                        kol_xlsx = build_shuidan_xlsx(kol_orders)
                        st.download_button(
                            f"🌟 下载深度达人水单（{len(kol_orders)} 单）",
                            data=kol_xlsx,
                            file_name=f"{date_compact}深度达人水单.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary", use_container_width=True, key="dl_kol",
                        )

                with col_other:
                    st.markdown("#### 📦 客户水单")
                    if len(customer_orders) == 0:
                        st.info("今日无客户单")
                    else:
                        other_xlsx = build_shuidan_xlsx(customer_orders)
                        st.download_button(
                            f"📦 下载客户水单（{len(customer_orders)} 单）",
                            data=other_xlsx,
                            file_name=f"{date_compact}客人水单.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary", use_container_width=True, key="dl_other",
                        )

                # —— Return Label 单独区域 ——
                st.divider()
                st.markdown("### ↩️ Return Label 包裹（顾客/达人寄回）")
                st.caption(
                    "**这些是额外生成的寄回包裹 label**：发件人 = 顾客/达人，收件人 = NailVesta。"
                    "Return Label 打勾的订单仍然已经包含在上面的正常发货水单里；这里是额外再生成一张寄回 label ⚠️"
                )

                if return_total == 0:
                    st.info("今日无 Return Label 包裹 ✅")
                else:
                    col_ret_kol, col_ret_cust = st.columns(2)

                    with col_ret_kol:
                        st.markdown("#### 🌟 退货 · 深度达人")
                        if len(return_kol_orders) == 0:
                            st.info("今日无深度达人退货")
                        else:
                            with st.expander(f"👀 查看 {len(return_kol_orders)} 单详情", expanded=False):
                                ret_kol_cols = ["Order ID", "客诉类型", "款式", "Handle", "达人Name", "地址"]
                                ret_kol_preview = return_kol_orders[
                                    [c for c in ret_kol_cols if c in return_kol_orders.columns]
                                ]
                                st.dataframe(ret_kol_preview, use_container_width=True, height=200)
                            ret_kol_xlsx = build_return_label_xlsx(return_kol_orders)
                            st.download_button(
                                f"↩️ 下载深度达人退货水单（{len(return_kol_orders)} 单）",
                                data=ret_kol_xlsx,
                                file_name=f"{date_compact}Return_Label_深度达人水单.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary", use_container_width=True, key="dl_return_kol",
                            )

                    with col_ret_cust:
                        st.markdown("#### 📦 退货 · 客户")
                        if len(return_cust_orders) == 0:
                            st.info("今日无客户退货")
                        else:
                            with st.expander(f"👀 查看 {len(return_cust_orders)} 单详情", expanded=False):
                                ret_cust_cols = ["Order ID", "客诉类型", "Product Name", "Shipping Info"]
                                ret_cust_preview = return_cust_orders[
                                    [c for c in ret_cust_cols if c in return_cust_orders.columns]
                                ]
                                st.dataframe(ret_cust_preview, use_container_width=True, height=200)
                            ret_cust_xlsx = build_return_label_xlsx(return_cust_orders)
                            st.download_button(
                                f"↩️ 下载客户退货水单（{len(return_cust_orders)} 单）",
                                data=ret_cust_xlsx,
                                file_name=f"{date_compact}Return_Label_客人水单.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary", use_container_width=True, key="dl_return_cust",
                            )

                st.caption(
                    "💡 已强制邮编/电话/Order ID 为文本格式，"
                    "Excel 打开不会丢失前导零或 Order ID 转科学计数法"
                )

    # ============================================================
    # 深达水单（USPS 批量导入 CSV）—— 独立可选通道，不上传不报错
    # ============================================================
    st.divider()
    st.markdown("### 🌟 深达水单（USPS 批量导入格式，可选）")
    st.caption(
        "上传 Lark 深达水单 CSV → 选日期 → 下载 USPS 模板格式的深达水单 CSV。"
        "不上传这里不影响上面的客人水单生成。"
    )

    lark_file_kol = st.file_uploader(
        "Lark 深达水单 CSV（不上传则跳过）", type=["csv"], key="lark_kol"
    )

    if lark_file_kol is not None:
        df_kol = load_lark_data(lark_file_kol.read())
        warn_if_corrupted_order_ids(df_kol)
        all_dates_kol = sorted(
            [d for d in df_kol["日期"].dropna().unique()
             if re.match(r"^\d{4}/\d{1,2}/\d{1,2}$", str(d))],
            key=lambda x: datetime.strptime(x, "%Y/%m/%d"), reverse=True,
        )
        if not all_dates_kol:
            st.error("深达 CSV 中找不到合法日期")
        else:
            sel_date_kol = st.selectbox(
                "深达发货日期（默认最晚）",
                options=all_dates_kol, index=0,
                format_func=lambda x: f"{x} {'⬅️ 最新' if x == all_dates_kol[0] else ''}",
                key="d_kol",
            )
            orders_kol_all = filter_orders_for_date(df_kol, sel_date_kol)
            orders_kol = orders_kol_all[orders_kol_all.apply(is_kol_order, axis=1)]

            st.metric("当日深达单", len(orders_kol))

            if len(orders_kol) == 0:
                st.warning(f"⚠️ {sel_date_kol} 没有深达订单")
            else:
                kol_issues = build_address_issue_report(orders_kol, label_kind="深达 USPS")
                if len(kol_issues) > 0:
                    serious_cnt = (kol_issues["问题等级"] == "严重").sum()
                    remind_cnt = (kol_issues["问题等级"] == "提醒").sum()
                    st.warning(
                        f"⚠️ 地址信息检查发现 {len(kol_issues)} 条问题："
                        f"{serious_cnt} 条严重 / {remind_cnt} 条提醒。"
                    )
                    with st.expander("🚨 查看深达地址问题明细", expanded=True):
                        show_cols = [
                            "影响Label", "订单号", "客诉类型", "问题等级", "问题字段", "问题说明",
                            "解析姓名", "解析电话", "解析地址一", "解析地址二",
                            "解析城市", "解析州", "解析邮编", "原始地址信息",
                        ]
                        st.dataframe(
                            kol_issues[[c for c in show_cols if c in kol_issues.columns]],
                            use_container_width=True, height=320, hide_index=True,
                        )
                else:
                    st.success("✅ 深达地址信息检查通过。")

                date_compact_kol = sel_date_kol.replace("/", "")
                kol_usps_csv = build_kol_usps_csv(orders_kol, sel_date_kol)
                st.download_button(
                    f"🌟 下载深达水单 CSV（{len(orders_kol)} 单，USPS 格式）",
                    data=kol_usps_csv,
                    file_name=f"{date_compact_kol}深达水单_USPS.csv",
                    mime="text/csv",
                    type="primary", use_container_width=True, key="dl_kol_usps",
                )


# ============================================================
# Tab 2: 拣货单 & 发货单生成器（原独立程序全功能移植）
# ============================================================
with tab2:
    st.subheader("第 2 步 · 拣货单 & 发货单生成器")
    st.caption(
        "上传当日的 客人水单 CSV / 深度达人单 CSV（可只传其一）→ "
        "下载拣货单 CSV（按库位汇总）+ 发货单 PDF（按 Tracking No. 一单一页）"
    )

    if not REPORTLAB_OK:
        st.error("缺少 reportlab 依赖：请在 requirements.txt 加入 reportlab 后重新部署")
        st.stop()

    col_c, col_i = st.columns(2)

    with col_c:
        customer_file = st.file_uploader(
            "上传客人水单 CSV", type=["csv"], key="customer_upload"
        )

    with col_i:
        influencer_file = st.file_uploader(
            "上传深度达人单 CSV", type=["csv"], key="influencer_upload"
        )

    no_loc_choice = st.radio(
        "无固定库位商品（Organizer Binder / Toolkits / Storage Box 等杂项）如何处理？",
        ["✅ 添加 —— 正常计入拣货单/发货单，库位标注为「无固定库位·人工核对」",
         "🚫 不添加 —— 从拣货单/发货单中剔除，但会列出清单提醒人工核对"],
        index=0,
        key="no_loc_choice",
    )
    include_no_location = no_loc_choice.startswith("✅")

    st.divider()

    if customer_file:
        df_c = read_file(customer_file)
        warn_if_corrupted_tracking(df_c)
        rows_c, skipped_c = parse_customer_rows(df_c, include_no_location)
        date_c = extract_date_prefix(df_c)
        st.markdown("## 客人水单")
        if skipped_c:
            st.warning(
                "以下商品无固定库位、已从拣货单/发货单中剔除，请人工核对是否需要打包：\n"
                + "\n".join(f"- {n}：{p}" for n, p in skipped_c)
            )
        show_section("客人", rows_c, "c", date_c)
        st.divider()

    if influencer_file:
        df_i = read_file(influencer_file)
        warn_if_corrupted_tracking(df_i)
        rows_i, skipped_i = parse_influencer_rows(df_i, include_no_location)
        date_i = extract_date_prefix(df_i)
        st.markdown("## 深度达人单")
        if skipped_i:
            st.warning(
                "以下商品无固定库位、已从拣货单/发货单中剔除，请人工核对是否需要打包：\n"
                + "\n".join(f"- {n}：{p}" for n, p in skipped_i)
            )
        show_section("达人", rows_i, "i", date_i)

    if not customer_file and not influencer_file:
        st.info("请上传客人水单或深度达人单 CSV 文件以生成报表。")
